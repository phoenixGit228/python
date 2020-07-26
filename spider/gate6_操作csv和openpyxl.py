#%%
file=open('test.csv','a+')
#创建test.csv文件，以追加的读写模式
file.write('美国队长,钢铁侠,蜘蛛侠')
#写入test.csv文件
file.close()
#关闭文件

# %%
import openpyxl
wb = openpyxl.workbook()

#%%
import openpyxl 
wb=openpyxl.Workbook() 
sheet=wb.active
sheet.title='new title'
sheet['A1'] = '漫威宇宙'
rows= [['美国队长','钢铁侠','蜘蛛侠'],['是','漫威','宇宙', '经典','人物']]
for i in rows:
    sheet.append(i)
print(rows)
wb.save('Marvel.xlsx')

# %%
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'new sheet'
sheet['a1'] = '漫威宇宙'
rows = [['美国队长','钢铁侠','黑寡妇'],['绿巨人','鹰眼']]
for i in rows:
    sheet.append(i)
wb.save('Marvel.xlsx')

#%%
wb = openpyxl.load_workbook('Marvel.xlsx')
sheet = wb['new sheet']
sheetname = wb.sheetnames
print(sheetname)
cell_a1 = sheet['a1']
cell_a1_value = cell_a1.value
print(cell_a1_value)
# %%
# 引用csv模块。
import csv

# 调用open()函数打开csv文件，传入参数：文件名“demo.csv”、写入模式“w”、newline=''、encoding='utf-8'。
csv_file = open('demo.csv','w',newline='',encoding='utf-8')
# 用csv.writer()函数创建一个writer对象。
writer = csv.writer(csv_file)
# 调用writer对象的writerow()方法，可以在csv文件里写入一行文字 “电影”和“豆瓣评分”。
writer.writerow(['电影','豆瓣评分'])
# 在csv文件里写入一行文字 “银河护卫队”和“8.0”。
writer.writerow(['银河护卫队','8.0'])
# 在csv文件里写入一行文字 “复仇者联盟”和“8.1”。
writer.writerow(['复仇者联盟','8.1'])
# 写入完成后，关闭文件就大功告成啦！
csv_file.close()


# %%
import csv
csv_file = open('rating.csv', 'w', newline='', encoding='utf-8')
writer = csv.writer(csv_file)
writer.writerow(['电影','豆瓣评分'])
writer.writerow(['银河护卫队',8.0])
writer.writerow(['复仇者联盟',8.1])
csv_file.close()

# %%
"""保存周杰伦歌曲，专辑，链接等等"""
import requests,openpyxl
# 创建工作簿
wb=openpyxl.Workbook()  
# 获取工作簿的活动表
sheet=wb.active 
# 工作表重命名
sheet.title='lyrics' 

sheet['A1'] ='歌曲名'     # 加表头，给A1单元格赋值
sheet['B1'] ='所属专辑'   # 加表头，给B1单元格赋值
sheet['C1'] ='播放时长'   # 加表头，给C1单元格赋值
sheet['D1'] ='播放链接'   # 加表头，给D1单元格赋值

url = 'https://c.y.qq.com/soso/fcgi-bin/client_search_cp'
for x in range(5):
    params = {
        'ct': '24',
        'qqmusic_ver': '1298',
        'new_json': '1',
        'remoteplace': 'txt.yqq.song',
        'searchid': '64405487069162918',
        't': '0',
        'aggr': '1',
        'cr': '1',
        'catZhida': '1',
        'lossless': '0',
        'flag_qc': '0',
        'p': str(x + 1),
        'n': '20',
        'w': '周杰伦',
        'g_tk': '5381',
        'loginUin': '0',
        'hostUin': '0',
        'format': 'json',
        'inCharset': 'utf8',
        'outCharset': 'utf-8',
        'notice': '0',
        'platform': 'yqq.json',
        'needNewCode': '0'
    }

    res_music = requests.get(url, params=params)
    json_music = res_music.json()
    list_music = json_music['data']['song']['list']
    for music in list_music:
        # 以name为键，查找歌曲名，把歌曲名赋值给name
        name = music['name']
        # 查找专辑名，把专辑名赋给album
        album = music['album']['name']
        # 查找播放时长，把时长赋值给time
        time = music['interval']
        # 查找播放链接，把链接赋值给link
        link = 'https://y.qq.com/n/yqq/song/' + str(music['mid']) + '.html\n\n'
        # 把name、album、time和link写成列表，用append函数多行写入Excel
        sheet.append([name,album,time,link])  
        print('歌曲名：' + name + '\n' + '所属专辑:' + album +'\n' + '播放时长:' + str(time) + '\n' + '播放链接:'+ link)
        
# 最后保存并命名这个Excel文件        
wb.save('Jay.xlsx') 

# %%
import requests,openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet_title = 'lyric'
sheet['a1'] = '歌名'
sheet['b1'] = '专辑'
sheet['c1'] = '时长'
sheet['d1'] = '链接'

url = 'https://c.y.qq.com/soso/fcgi-bin/client_search_cp'
for x in range(5):
    headers = {
        'origin':'https://y.qq.com',
        # 请求来源，本案例中其实是不需要加这个参数的，只是为了演示
        'referer':'https://y.qq.com/portal/search.html',
        # 请求来源，携带的信息比“origin”更丰富，本案例中其实是不需要加这个参数的，只是为了演示
        'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36',
        # 标记了请求从什么设备，什么浏览器上发出
    }
    params = {
        'ct': '24',
        'qqmusic_ver': '1298',
        'new_json': '1',
        'remoteplace': 'sizer.yqq.song_next',
        'searchid': '64405487069162918',
        't': '0',
        'aggr': '1',
        'cr': '1',
        'catZhida': '1',
        'lossless': '0',
        'flag_qc': '0',
        'p': str(x + 1),
        'n': '20',
        'w': '周杰伦',
        'g_tk': '5381',
        'loginUin': '0',
        'hostUin': '0',
        'format': 'json',
        'inCharset': 'utf8',
        'outCharset': 'utf-8',
        'notice': '0',
        'platform': 'yqq.json',
        'needNewCode': '0'
    }

    res_music = requests.get(url, headers=headers,params=params)
    json_music = res_music.json()
    list_music = json_music['data']['song']['list']
    for music in list_music:
        sheet.append([music['name'],music['album']['name'],music['interval'],'https://y.qq.com/n/yqq/song/'+music['file']['media_mid'] + '.html'])
        # print(music['name'])
        # print('所属专辑：' + music['album']['name'])
        # print('播放时长：' + str(music['interval']) + '秒')
        # print('播放链接：https://y.qq.com/n/yqq/song/' + music['file']['media_mid'] + '.html\n\n')
        
wb.save('jay_zhou.xlsx')

# %%
"""爬取豆瓣电影250，序号/电影名/评分/推荐语/链接 并保存到csv和xlsx文件中"""
import requests, bs4, csv, openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = '豆瓣电影'
sheet['a1'] = '序号'
sheet['b1'] = '电影名'
sheet['c1'] = '评分'
sheet['d1'] = '推荐语'
sheet['e1'] = '链接'

csv_file = open('douban_movie.csv','w',newline='',encoding='utf-8')
writer = csv.writer(csv_file)
writer.writerow(['豆瓣电影', '序号', '电影名', '评分', '推荐语', '链接'])

headers={'user-agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36'}
for x in range(10):
    url = 'https://movie.douban.com/top250?start=' + str(x*25) + '&filter='
    res = requests.get(url, headers=headers)
    bs = bs4.BeautifulSoup(res.text, 'html.parser')
    bs = bs.find('ol', class_="grid_view")
    for titles in bs.find_all('li'):
        num = titles.find('em',class_="").text
        title = titles.find('span', class_="title").text
        comment = titles.find('span',class_="rating_num").text
        url_movie = titles.find('a')['href']

        if titles.find('span',class_="inq") != None:
            tes = titles.find('span',class_="inq").text
            sheet.append([num, title, comment, tes, url_movie])
            writer.writerow([num, title, comment, tes, url_movie])
            # print(num + '.' + title + '——' + comment + '\n' + '推荐语：' + tes +'\n' + url_movie)
        else:
            sheet.append([num, title, comment, '', url_movie])
            writer.writerow([num, title, comment, '', url_movie])
            #print(num + '.' + title + '——' + comment + '\n' +'\n' + url_movie)
wb.save('douban_movie.xlsx')
csv_file.close()

#%%
